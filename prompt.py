from langchain.agents import initialize_agent, AgentType
from langchain_community.agent_toolkits.load_tools import load_tools
from langchain_community.llms import HuggingFacePipeline
from langchain_community.llms.baseten import LLM
from langchain_community.utilities import SerpAPIWrapper
from langchain_community.utilities import WikipediaAPIWrapper
from langchain_community.tools import WikipediaQueryRun
from transformers import AutoTokenizer, AutoModelForCausalLM
from typing import Dict, List, Optional, Tuple, Union
import openpyxl, tqdm, time, requests
from typing import Optional, List, Dict, Mapping, Any


import os, json, argparse, LLM_utils

parser = argparse.ArgumentParser()

parser.add_argument("--test_model", type=str, default='ChatGPT')
parser.add_argument("--file_path", type=str, default='X:/raw data-first step.xlsx')
parser.add_argument("--question_column", type=int, default=1)
parser.add_argument("--column", type=int, default=25)
args = parser.parse_args()

llm = eval("LLM_utils.{}".format(args.test_model))()

wb = openpyxl.load_workbook(args.file_path)
sheets = wb.sheetnames
for i in range(len(sheets)):
    w = wb[sheets[i]]
    # w.cell(row=1, column=12, value="Baichuan2-7B")
    for j in tqdm.tqdm(range(2, w.max_row + 1)):
        if w.cell(j, args.column).value != None:
            continue
        question:str= w.cell(j, args.question_column).value
        while True:
            response:str = llm(question)
            if response == "Request model!":
                time.sleep(10)
            else:
                break
        print(response)
        w.cell(row=j, column=args.column, value=response)
        wb.save(args.file_path)
